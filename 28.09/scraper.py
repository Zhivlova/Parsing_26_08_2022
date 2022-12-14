import os
from openpyxl import Workbook, load_workbook
import csv

import pandas as pd

mydir = '/Users/natalazivlova/Desktop/parser/28.09/'
myfile = 'рейтинг 2021 корр.xlsx'
file = os.path.join(mydir, myfile)

wb = load_workbook(file, data_only=True)

print(f'Структура эксель файла {wb.sheetnames}')
print('------------------------------------------------------------------------------')

sheet = wb.worksheets[0]

sonko = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    sonko['subject'].append(subject)
    scores = sheet[row][2].value
    sonko['scores'].append(scores)
sonko = pd.DataFrame(sonko)

social_entrepreneurship = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    social_entrepreneurship['subject'].append(subject)
    scores = sheet[row][3].value
    social_entrepreneurship['scores'].append(scores)
social_entrepreneurship = pd.DataFrame(social_entrepreneurship)

sonco_taxes = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    sonco_taxes['subject'].append(subject)
    scores = sheet[row][4].value
    sonco_taxes['scores'].append(scores)
sonco_taxes = pd.DataFrame(sonco_taxes)

taxes_donations = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    taxes_donations['subject'].append(subject)
    scores = sheet[row][5].value
    taxes_donations['scores'].append(scores)
taxes_donations = pd.DataFrame(taxes_donations)

mun_prog_sonko = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    mun_prog_sonko['subject'].append(subject)
    scores = sheet[row][6].value
    mun_prog_sonko['scores'].append(scores)
print(mun_prog_sonko)
mun_prog_sonko = pd.DataFrame(mun_prog_sonko)

mun_prog_sopred = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    mun_prog_sopred['subject'].append(subject)
    scores = sheet[row][7].value
    mun_prog_sopred['scores'].append(scores)
mun_prog_sopred = pd.DataFrame(mun_prog_sopred)

infrastructure = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    infrastructure['subject'].append(subject)
    scores = sheet[row][8].value
    infrastructure['scores'].append(scores)
infrastructure = pd.DataFrame(infrastructure)

workers = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    workers['subject'].append(subject)
    scores = sheet[row][9].value
    workers['scores'].append(scores)
workers = pd.DataFrame(workers)

social_entrepreneurship_support = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    social_entrepreneurship_support['subject'].append(subject)
    scores = sheet[row][10].value
    social_entrepreneurship_support['scores'].append(scores)
social_entrepreneurship_support = pd.DataFrame(social_entrepreneurship_support)

soc_order = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    soc_order['subject'].append(subject)
    scores = sheet[row][11].value
    soc_order['scores'].append(scores)
soc_order = pd.DataFrame(soc_order)

personal_financing = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    personal_financing['subject'].append(subject)
    scores = sheet[row][12].value
    personal_financing['scores'].append(scores)
personal_financing = pd.DataFrame(personal_financing)

gchp_not_public = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    gchp_not_public['subject'].append(subject)
    scores = sheet[row][13].value
    gchp_not_public['scores'].append(scores)
gchp_not_public = pd.DataFrame(gchp_not_public)

website = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    website['subject'].append(subject)
    scores = sheet[row][14].value
    website['scores'].append(scores)
website = pd.DataFrame(website)

clicks = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    clicks['subject'].append(subject)
    scores = sheet[row][15].value
    clicks['scores'].append(scores)
clicks = pd.DataFrame(clicks)

public_licenses = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    public_licenses['subject'].append(subject)
    scores = sheet[row][16].value
    public_licenses['scores'].append(scores)
public_licenses = pd.DataFrame(public_licenses)

open_data = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    open_data['subject'].append(subject)
    scores = sheet[row][17].value
    open_data['scores'].append(scores)
open_data = pd.DataFrame(open_data)

children = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    children['subject'].append(subject)
    scores = sheet[row][18].value
    children['scores'].append(scores)
children = pd.DataFrame(children)

medical_organizations = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    medical_organizations['subject'].append(subject)
    scores = sheet[row][19].value
    medical_organizations['scores'].append(scores)
medical_organizations = pd.DataFrame(medical_organizations)

suppliers = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    suppliers['subject'].append(subject)
    scores = sheet[row][20].value
    suppliers['scores'].append(scores)
suppliers = pd.DataFrame(suppliers)

culture = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    culture['subject'].append(subject)
    scores = sheet[row][21].value
    culture['scores'].append(scores)
culture = pd.DataFrame(culture)

final_results = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    final_results['subject'].append(subject)
    scores = sheet[row][22].value
    final_results['scores'].append(scores)
final_results = pd.DataFrame(final_results)



with open('scores.csv', 'w') as file:
    writer = csv.writer(file)
    writer.writerow(["Субъект", "Баллы"])
# for i in final_results:
#     subject = i[0]
#     scores = i[1]
#
#     with open('scores.csv', 'w') as file:
#         writer = csv.writer(file)
#         writer.writerow(
#             i
#         )